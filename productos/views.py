from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http.response import JsonResponse
from .models import Producto, Categoria
import csv
from django.core.files.storage import FileSystemStorage
from .functions import *
from django.contrib import messages
import openpyxl
from io import BytesIO
from django.http import HttpResponse
# Create your views here.
app_name = 'productos'

# Muestra la tabla de productos.


@login_required
def productos(request):
    autorizado = request.session.get('autorizado')
    usuario_nombre = request.session.get('usuario_nombre')
    img = request.session.get('img')
    prods = Producto.objects.filter(resultado=0).all()
    if request.user.is_superuser or request.user.groups.filter(name='Gerencia').exists():
        autorizado = True
    data = {
        'usuario': usuario_nombre,
        'img': img,
        'prods': prods,
        'autorizado': autorizado,
    }
    return render(request, 'productos/productos.html', data)

# Vista que se usa a traves de una petición AJAX desde el frontend para obtener la información de los usuarios y de los proyectos en formato JSON.
# y utilizarla para completar un modal de edición de productos.


@login_required
def info_editar_producto(request):
    Cate = Categoria.objects.all()

    Cate_data = list(Cate.values('nombre'))
    data = {
        'Cate': Cate_data,
    }
    return JsonResponse(data)

# Vista que recibe la información de los modales de editar y agregar producto para luego actualizar la base de datos.


@login_required
def editar_producto(request):
    if request.method == 'POST':
        if request.POST['codigo'] != '':
            cambios = False
            try:
                info = "|".join(request.POST.values())
                info = info.split('|')
                p = Producto.objects.get(pk=request.POST['codigo'])
                if p.nombre != info[2]:
                    p.nombre = info[2]
                    cambios = True
                if p.ancho != info[5]:
                    p.ancho = info[5].replace(',', '.')
                    cambios = True
                if p.alto != info[6]:
                    p.alto = info[6].replace(',', '.')
                    cambios = True
                if p.precio != info[3]:
                    p.precio = info[3].replace(',', '.')
                    cambios = True
                if p.categoria.nombre != info[4]:
                    p.categoria = Categoria.objects.get(nombre=info[4])
                    cambios = True
                if p.factor != info[7]:
                    p.factor = info[7].replace(',', '.')
                    cambios = True
                if cambios:
                    p.save()
                    messages.success(
                        request, 'Los datos del producto se han actualizado correctamente.')
            except Exception as e:
                messages.error(
                    request, f'No se ha podido actualizar los datos del producto. Error({e})')
        else:
            try:
                p = Producto.objects.create(
                    presupuesto=None,
                    nombre=request.POST['nombre_add'],
                    ancho=request.POST['ancho_add'],
                    alto=request.POST['alto_add'],
                    precio=request.POST['precio_add'],
                    categoria=Categoria.objects.get(
                        nombre=request.POST['categ_add']),
                    factor=request.POST['factor_add'],
                )
                messages.success(
                    request, 'El nuevo producto se ha agregado correctamente.')
            except Exception as e:
                messages.error(
                    request, f'No se ha podido agregar el nuevo producto. Error({e})')
    return redirect('/productos', messages)

# Vista que permite borrar un producto de la base de datos.


@login_required
def borrar_producto(request, producto_id):
    try:
        prod = Producto.objects.filter(pk=producto_id)
        prod.delete()
        messages.success(request, 'El producto se ha borrado correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se ha podido borrar el producto. Error({e})')
    return redirect('/Productos', messages)

# Función que carga productos desde un archivo excel.


def cargar_productos(request):
    total = 0
    contador = 0
    try:
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                total += 1
                if total == 1:
                    continue

                # Verificar si la fila está vacía
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    break  # Detener el procesamiento si la fila está vacía

                values = []
                for col in range(1, 8):
                    cell = sheet.cell(row=total, column=col)
                    value = cell.value if cell.value is not None else " "
                    values.append(str(value))
                codigo, nombre, ancho, alto, precio, factor, categoria = values
                if not comparar(codigo):
                    try:
                        Producto.objects.create(
                            codigo=codigo,
                            nombre=nombre,
                            precio=precio,
                            factor=factor,
                            ancho=ancho if ancho != ' ' else 1,
                            alto=alto if alto != ' ' else 1,
                            categoria=Categoria.objects.get(
                                nombre=categoria),
                        )
                        contador += 1
                    except Exception as e:
                        messages.error(
                            request, f"Sucedió un error inesperado. Error({e})")
            messages.success(
                request, f'Se han cargado {contador} de {total} productos correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se han podido cargar los productos. Error({e})')
    return redirect('/productos', messages)


# Vista que permite general un excel con los productos de la base de datos.
@login_required
def exportar_productos(request):
    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Backup_Lista de productos'

    # Definir encabezados
    headers = ['Código', 'Nombre', 'Ancho',
               'Alto', 'Precio', 'Factor', 'Categoría']
    sheet.append(headers)

    # Agregar datos de los productos
    productos = Producto.objects.all().filter(presupuesto=None)
    for producto in productos:
        data = [
            producto.codigo,
            producto.nombre,
            producto.ancho,
            producto.alto,
            producto.precio,
            producto.factor,
            producto.categoria.nombre
        ]
        sheet.append(data)

    # Guardar el libro en un buffer de memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Crear la respuesta HTTP
    response = HttpResponse(
        buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    return response

# Vista que permite borrar todos los productos.


def borrar_todos_productos(request):
    try:
        Producto.objects.all().filter(presupuesto=None).filter(vendedor=None).delete()
        messages.success(
            request, 'Todos los productos se han borrado correctamente.')
    except Exception as e:
        messages.error(
            request, f'No se han podido borrar todos los productos. Error({e})')
    return redirect('/productos', messages)
